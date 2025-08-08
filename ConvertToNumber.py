import os
import sys
from openpyxl import load_workbook # type: ignore
from openpyxl.styles import numbers #type: ignore
from datetime import datetime
import tkinter as tk
from tkinter import ttk
from pathlib import Path

pathex=[os.path.abspath('.')]
root = tk.Tk()
status_label = tk.Label(root, text="", fg="green")
status_label.pack(side='top', fill='x')
#Custom UI Functions
def populate_tree(tree, node, path):
    tree.delete(*tree.get_children(node))
    try:
        for name in os.listdir(path):
            full_path = os.path.join(path, name)
            if os.path.isdir(full_path):
                child = tree.insert(node, 'end', text=name, open=False, values=[full_path])
                tree.insert(child, 'end')
    except PermissionError:
        pass

def get_full_path(node):
    return tree.item(node, 'values')[0]

def on_tree_expand(event):
    node = tree.focus()
    path = get_full_path(node)
    tree.delete(*tree.get_children(node))
    try:
        for name in os.listdir(path):
            full_path = os.path.join(path, name)
            if os.path.isdir(full_path):
                child = tree.insert(node, 'end', text=name, values=[full_path])
                tree.insert(child, 'end')
        show_files(path)
    except PermissionError:
        pass

def on_tree_select(event):
    node = tree.focus()
    path = get_full_path(node)
    show_files(path)
    tar.set(path)  # Update selected path
    status_label.config(text = "")
    
def show_files(folder_path):
    file_listbox.delete(0, tk.END)
    try:
        for f in os.listdir(folder_path):
            full_path = os.path.join(folder_path, f)
            if os.path.isfile(full_path):
                file_listbox.insert(tk.END, f)
        tar.set(folder_path)
    except Exception as e:
        file_listbox.insert(tk.END, f"Error: {e}")

def convert_string_numbers_to_numeric(targetFile, outPath):
    try:
        wb = load_workbook(targetFile)
        #Iterate spreadsheets
        for ws in wb.worksheets:
            #Iterate rows
            for r in ws.iter_rows():
                #Iterate cells
                for c in r:
                    if isinstance(c.value, str) and c.value.strip().replace('.', '', 1).isdigit():
                        try:
                            stripped = c.value.strip()
                            if '.' in stripped:
                                num = float(c.value)
                                c.value = num
                            else:
                                num = float(c.value)
                                c.value = int(num)
                                c.number_format = numbers.FORMAT_NUMBER
                        except ValueError:
                            pass
        wb.save(targetFile)
        fileLog = os.path.join(os.path.dirname(targetFile), outPath)
        with open(fileLog,'a') as output:
            output.write(f"Processed: {targetFile}\n")
    except Exception as e:
        print(f"Failed to Process {targetFile}: {e}\n")

#Log run when folder found
def folder_find():
    if tar.get():
        try:
            curr_datetime = datetime.now()
            formatted = curr_datetime.strftime("%Y-%m-%d %H:%M")
            target_folder = tar.get()
            folder_n = os.path.basename(target_folder)
            #Edited
            outputLog = f"{folder_n}_CtN_output.log"
            fileLog = os.path.join(target_folder,outputLog)

            with open(fileLog, 'a') as output:
                output.write(f"Timestamp: {formatted}\n")
            # Process all Excel files in the target folder
            for filename in os.listdir(tar.get()):
                targetFile = os.path.join(tar.get(), filename)

                if filename == outputLog:
                    continue

                if filename.lower().endswith((".xlsx",".xlsm",".xltx",".xltm")):
                    convert_string_numbers_to_numeric(targetFile, fileLog)
                elif os.path.isdir(targetFile) is False:
                    with open(fileLog, 'a') as output:
                        output.write(f"Not Processed: {targetFile}\n")
            with open(fileLog,'a') as output:
                output.write("\n\n\n")
            status_label.config(fg = 'green',text=f"Conversion report has been logged at {tar.get()}/{folder_n}_CtN_output.log")
        except Exception as e:
            if ".log" in str(e):
                status_label.config(fg= 'red', text = f"Error: {e}, Unable to Create Output Log in Specified Directory. Please Try Another Folder")
            else:
                status_label.config(fg= 'red', text = f"Error: {e}, Please Try Another Folder")

if __name__ == "__main__":
    if getattr(sys, 'frozen', False):
        # .exe
        origin = os.path.dirname(sys.executable)
    else:
        # .py
        origin = os.path.dirname(os.path.abspath(__file__))
    outPath = "CtN_Output.log"
    # UI setup
root.title("ConvertToNumber")
root.geometry("600x400")

tar = tk.StringVar()

# Left-side frame
left_frame = tk.Frame(root)
left_frame.pack(side='left', fill='y', padx=5)

folder_header = ttk.Label(left_frame, text="Folders", anchor='w', padding=5)
folder_header.pack(fill='x')

# Treeview for folder navigation
tree = ttk.Treeview(left_frame, show='tree')
tree.pack(fill='y', expand=True)

# Populate root directories
if os.name == 'nt':
    user_profile = Path.home()
    special_fold = {
        "OneDrive": user_profile / "OneDrive",
        "OneDrive - Diasorin-Luminex": user_profile / "OneDrive - Diasorin-Luminex"
    }
    for name, path in special_fold.items():
        if path.exists():
            node = tree.insert('', 'end', text=name, open=False, values=[str(path)])
            tree.insert(node, 'end')

    drives = [f"{d}:\\" for d in "ABCDEFGHIJKLMNOPQRSTUVWXYZ" if os.path.exists(f"{d}:\\")]
    for drive in drives:
        node = tree.insert('', 'end', text=drive, open=False, values=[drive])
        tree.insert(node, 'end')
else:
    # macOS/Linux: start from root "/"
    root_path = "/"
    node = tree.insert('', 'end', text=root_path, open=False, values=[root_path])
    tree.insert(node, 'end')

tree.bind('<<TreeviewOpen>>', on_tree_expand)
tree.bind('<<TreeviewSelect>>', on_tree_select)

# Right-side frame
right_frame = tk.Frame(root)
right_frame.pack(side='right', fill='both', expand=True)

file_header = ttk.Label(right_frame, text="File Preview", anchor='w', padding=5)
file_header.pack(fill='x')

# Preview Files
file_listbox = tk.Listbox(right_frame, bg="white")
file_listbox.pack(fill='both', expand=True)

# Confirm button
confirm_button = tk.Button(folder_header, text="Start Conversion", command=folder_find)
confirm_button.pack(side = 'right')
confirm_button.configure(background="Light green")

root.mainloop()